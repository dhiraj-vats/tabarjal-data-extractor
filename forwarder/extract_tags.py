"""
Read RequiredDocument.xlsx, pull every green-highlighted tag from the PQM,
WMS, and SACU sheets, and emit tags.json — the runtime tag manifest used by
the forwarder. Run once whenever the Excel changes.

Usage:
    python extract_tags.py [path/to/RequiredDocument.xlsx] [-o tags.json]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

GREEN_RGBS = {"FF00FF00", "0000FF00", "FF00B050", "FF92D050"}

PQM_SOURCE_RE = re.compile(r"^PQM\s*-?\s*0?(\d+)$", re.I)
SACU_SHEET_RE = re.compile(r"^SACU\s+(MVPS\d+)$", re.I)

# Schema-defined component_key for each green tag. Lookup is by the
# unprefixed remainder of the Tagname (everything after the source code).
PQM_COMPONENT_BY_LABEL = {
    "TOTAL ACTIVE POWER": "total_active_power",
}

WMS_COMPONENT_BY_LABEL = {
    "DNI WM2": "dni_wm2",
    "FRONT SOIL SENSOR 1": "front_soil_sensor_1",
    "FRONT SOIL SENSOR 2": "front_soil_sensor_2",
    "FRONT TR LOSS SENSOR 1": "front_tr_loss_sensor_1",
    "FRONT TR LOSS SENSOR 2": "front_tr_loss_sensor_2",
    "GHI W": "ghi_w",
    "MODULE TEMPERATURE 1": "module_temperature_1",
    "POA1 W": "poa1_w",
    "POA2 W": "poa2_w",
    "WIND DIRECTION": "wind_direction",
    "WIND SPEED": "wind_speed",
}

SACU_COMPONENT_BY_LABEL = {
    "SACU DC CURR": "sacu_dc_curr",
    "SACU ACTIVE POWER": "sacu_active_power",
    "SACU PLANT STATUS 2": "sacu_plant_status_2",
    "SACU INV EFFICIENCY": "sacu_inv_efficiency",
}

IP_PORT_RE = re.compile(r"IP:\s*([\d.]+)\s*Port:\s*(\d+)", re.I)


def is_green(cell) -> bool:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    fg = fill.fgColor
    if fg is None or fg.type != "rgb" or fg.rgb is None:
        return False
    return fg.rgb.upper() in GREEN_RGBS


def parse_endpoint(raw: str | None) -> str | None:
    if not raw:
        return None
    m = IP_PORT_RE.search(str(raw))
    if not m:
        return None
    return f"opc.tcp://{m.group(1)}:{m.group(2)}"


def _section_header_for(ws, row_idx: int) -> tuple[int, str | None]:
    """Walk upward from row_idx to find the section header cell in column A.

    Returns (header_row, header_value). The section header is the most-recent
    column-A cell that has a value but is not the literal "S.No" column header.
    """
    for r in range(row_idx, 0, -1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        s = str(v).strip()
        if s.upper() == "S.NO":
            continue
        if s.replace(".", "").replace("0", "").strip() == "":
            continue
        try:
            float(s)
            continue
        except ValueError:
            pass
        return r, s
    return 0, None


def _endpoint_for_section(ws, header_row: int) -> str | None:
    for r in range(header_row, min(header_row + 5, ws.max_row + 1)):
        ep = parse_endpoint(ws.cell(row=r, column=4).value)
        if ep:
            return ep
    return None


def extract_pqm(ws) -> list[dict]:
    out: list[dict] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.column != 2 or not is_green(cell):
                continue
            tag = str(cell.value).strip()
            opc = ws.cell(row=cell.row, column=3).value
            opc = str(opc).strip() if opc else None
            header_row, header_val = _section_header_for(ws, cell.row)
            endpoint = _endpoint_for_section(ws, header_row)
            # Source code: turn "PQM -01" / "PQM - 02" into "PQM01" / "PQM02"
            m = PQM_SOURCE_RE.match((header_val or "").replace(" ", ""))
            if not m:
                raise RuntimeError(
                    f"PQM section header not recognised at row {header_row}: {header_val!r}"
                )
            source_code = f"PQM{int(m.group(1)):02d}"
            label = tag[len(source_code):].strip()
            component_key = PQM_COMPONENT_BY_LABEL.get(label)
            if component_key is None:
                raise RuntimeError(
                    f"No component_key mapping for PQM tag {tag!r} (label={label!r})"
                )
            out.append({
                "category": "PQM",
                "source_code": source_code,
                "component_key": component_key,
                "tagname": tag,
                "opc_node_id": opc,
                "opc_endpoint": endpoint,
            })
    return out


def extract_wms(ws) -> list[dict]:
    out: list[dict] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.column != 2 or not is_green(cell):
                continue
            tag = str(cell.value).strip()
            opc = ws.cell(row=cell.row, column=3).value
            opc = str(opc).strip() if opc else None
            header_row, header_val = _section_header_for(ws, cell.row)
            endpoint = _endpoint_for_section(ws, header_row)
            source_code = (header_val or "").strip()  # e.g. MVPS02
            label = tag[len(source_code):].strip()
            component_key = WMS_COMPONENT_BY_LABEL.get(label)
            if component_key is None:
                raise RuntimeError(
                    f"No component_key mapping for WMS tag {tag!r} (label={label!r})"
                )
            out.append({
                "category": "WMS",
                "source_code": source_code,
                "component_key": component_key,
                "tagname": tag,
                "opc_node_id": opc,
                "opc_endpoint": endpoint,
            })
    return out


def extract_sacu(ws, sheet_name: str) -> list[dict]:
    m = SACU_SHEET_RE.match(sheet_name)
    if not m:
        return []
    source_code = m.group(1).upper()  # e.g. MVPS02
    endpoint = None
    for r in range(1, min(ws.max_row + 1, 6)):
        endpoint = parse_endpoint(ws.cell(row=r, column=4).value)
        if endpoint:
            break
    out: list[dict] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.column != 2 or not is_green(cell):
                continue
            tag = str(cell.value).strip()
            opc = ws.cell(row=cell.row, column=3).value
            opc = str(opc).strip() if opc else None
            label = tag[len(source_code):].strip()
            component_key = SACU_COMPONENT_BY_LABEL.get(label)
            if component_key is None:
                raise RuntimeError(
                    f"No component_key mapping for SACU tag {tag!r} (label={label!r})"
                )
            out.append({
                "category": "SACU",
                "source_code": source_code,
                "component_key": component_key,
                "tagname": tag,
                "opc_node_id": opc,
                "opc_endpoint": endpoint,
            })
    return out


def extract(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    tags: list[dict] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        if sn == "PQM":
            tags.extend(extract_pqm(ws))
        elif sn == "WMS":
            tags.extend(extract_wms(ws))
        elif sn.startswith("SACU"):
            tags.extend(extract_sacu(ws, sn))
    return tags


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "xlsx",
        nargs="?",
        default=str(Path(__file__).resolve().parent.parent / "RequiredDocument .xlsx"),
        help="Path to the Excel sheet (default: ../RequiredDocument .xlsx)",
    )
    ap.add_argument(
        "-o", "--output",
        default=str(Path(__file__).resolve().parent / "tags.json"),
        help="Where to write the tag manifest (default: ./tags.json)",
    )
    args = ap.parse_args(argv)

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"error: {xlsx_path} not found", file=sys.stderr)
        return 1

    tags = extract(xlsx_path)
    missing_node = [t for t in tags if not t["opc_node_id"]]
    missing_ep = [t for t in tags if not t["opc_endpoint"]]
    if missing_node:
        print(f"error: {len(missing_node)} tags missing OPC node id", file=sys.stderr)
        return 2
    if missing_ep:
        print(f"error: {len(missing_ep)} tags missing OPC endpoint", file=sys.stderr)
        return 2

    Path(args.output).write_text(json.dumps(tags, indent=2) + "\n")

    by_cat: dict[str, int] = {}
    by_ep: dict[str, int] = {}
    for t in tags:
        by_cat[t["category"]] = by_cat.get(t["category"], 0) + 1
        by_ep[t["opc_endpoint"]] = by_ep.get(t["opc_endpoint"], 0) + 1
    print(f"wrote {len(tags)} tags to {args.output}")
    for cat, n in sorted(by_cat.items()):
        print(f"  {cat}: {n}")
    print("endpoints:")
    for ep, n in sorted(by_ep.items()):
        print(f"  {ep}: {n} tags")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
